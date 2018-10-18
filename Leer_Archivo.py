# -*- coding: utf-8 -*-

import os.path
import openpyxl
import shutil, os
import DataBase

def existencia_archivo(a):
    if os.path.exists(a):
        return True
    else:
        return False

def verificacion_columnas_base_data(a):
    documento  = openpyxl.load_workbook(a)
    hojas=documento.get_sheet_names()
    verif=0
    errores_columnas=[]
    for x in range(0,len(hojas)):
        if(hojas[x]=='Global Info'):
            verif=1
    if(verif==1):
        sheet = documento.get_sheet_by_name('Global Info')
        if(sheet['A1'].value!='CrmId'):
            errores_columnas.append('A1, NIT\n')
        if(sheet['B1'].value!='Pincode'):
            errores_columnas.append('B1, ID CLIENTE ONYX\n')
        if(sheet['C1'].value!='Fin'):
            errores_columnas.append('C1, NOMBRE CLIENTE\n')
        if(sheet['D1'].value!='sabor'):
            errores_columnas.append('D1, GRUPO OBJETIVO\n')
        return errores_columnas
    else:
        errores_columnas.append(0)
        return errores_columnas


def mover_archivo(origen,destino):
    if os.path.exists(origen):
        ruta = shutil.move(origen, destino)
        print('El directorio ha sido movido a', destino)
    else:
        print('El directorio origen no existe')

def cambiar_nombre_archivo(origen,nuevo_nombre):
    if os.path.exists(origen):
        os.rename(origen, nuevo_nombre)
    else:
        print('El directorio origen no existe')

def lectura_base_data(a):
    documento = openpyxl.load_workbook(a)
    sheet = documento.get_sheet_by_name('Global Info')
    registro=[]
    x=2
    #DataBase.truncate_base_tickets_stage()
    while(sheet.cell(row = x, column = 1).value != None):
        for y in range(1,5):
            value = sheet.cell(row = x, column = y).value
            registro.append(value)
        DataBase.insertar_registros(registro)
        del registro[:]
        x=x+1


#def lectura_base_dataI(a):
    #documento = openpyxl.load_workbook(a)
    #sheet = documento.get_sheet_by_name('Base Tickets')
    #registro=[]
    #x=2
    #DataBase.truncate_base_tickets_stage()
    #while(sheet.cell(row = x, column = 1).value != None):
        #for y in range(1,55):
            #value = sheet.cell(row = x, column = y).value
            #registro.append(value)
        #DataBase.insertar_registros(registro)
        #del registro[:]
        #x=x+1
    #DataBase.ejecutarSPCarga()